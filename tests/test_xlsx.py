from openpyxl import load_workbook
import os
from os_path.os_path_scripts import resources
xlsx_path = os.path.join(resources, 'file_example_XLSX_50.xlsx')


# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_xlsx_file():
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'
    assert sheet.cell(row=6, column=2).value == 'Nereida'
    assert sheet.cell(row=1, column=6).value == 'Age'
